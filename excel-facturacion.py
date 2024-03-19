import pandas as pd
import sqlite3
from openpyxl import load_workbook
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta

# Define the path to your Excel file
excel_file_path = r'C:\Users\tdiloreto\Algeiba\InO Billings - Documents\AIT_Billing_General_v2.xlsb.xlsx'

# Connect to SQLite Database
conn = sqlite3.connect('jiradatabase.db')

# Load the workbook
wb = load_workbook(excel_file_path, data_only=True)

# Load specific sheet into SQLite
sheet_name = 'Fact_Facturacion'
if sheet_name in wb.sheetnames:
    # Load data from the Excel file into a DataFrame
    data = pd.read_excel(excel_file_path, sheet_name=sheet_name, skiprows=11, engine='openpyxl')
    data.dropna(how='all', inplace=True)
    
    # Correctly reference columns in 'data' DataFrame to convert them into datetime
    date_columns = ['Fecha PF', 'Mes a facturar', 'Fecha Ren Tar', 'Fecha ren cont']
    for col in date_columns:
        if col in data.columns:
            data[col] = pd.to_datetime(data[col]).dt.strftime('%Y-%m-%d')

    # Write the DataFrame to the SQLite database, replacing the table if it exists
    data.to_sql('fact_facturacion', conn, if_exists='replace', index=False)

# Close the connection
conn.close()




# import pandas as pd
# import sqlite3
# from pyxlsb import open_workbook as open_xlsb

# # Define the path to your Excel file
# excel_file_path = r'C:\Users\tdiloreto\Algeiba\InO Billings - Documents\AIT_Billing_General_v2.xlsb.xlsx'

# # Connect to SQLite Database
# conn = sqlite3.connect('jiradatabase.db')

# # Read the specified sheet, skip rows, and use the 11th row as header
# df = pd.read_excel(excel_file_path, sheet_name='Fact_Facturacion', skiprows=10, engine='openpyxl')

# # Remove any completely empty rows that might still be present after the headers
# df.dropna(how='all', inplace=True)

# # Write the DataFrame to SQLite, using the sheet name as the table name
# df.to_sql('Fact_Facturacion', conn, if_exists='replace', index=False)

# # Close the connection
# conn.close()




# import pandas as pd
# import sqlite3
# from pyxlsb import open_workbook as open_xlsb

# # Define the path to your Excel file
# excel_file_path = r'C:\Users\tdiloreto\Algeiba\InO Billings - Documents\AIT_Billing_General_v2.xlsb.xlsx'

# # Connect to SQLite Database
# conn = sqlite3.connect('jiradatabase.db')

# # Use pyxlsb to read the specific table by name
# with open_xlsb(excel_file_path) as wb:
#     with wb.get_sheet('Fact_Facturacion') as sheet:
#         rows_gen = sheet.rows(skip=10)  # Skip the first 10 rows
#         headers = next(rows_gen)  # Get the 11th row as headers
#         headers = [h.v for h in headers]  # Extract the header values

#         # Convert rows into a list of dictionaries to use as DataFrame data
#         data = []
#         for row in rows_gen:
#             values = [r.v for r in row]  # Extract the cell values
#             if any(values):  # Only add rows that have any non-empty value
#                 data.append(dict(zip(headers, values)))

# # Create a DataFrame
# df = pd.DataFrame(data)

# # Write the DataFrame to SQLite, using the sheet name as the table name
# df.to_sql('Fact_Facturacion', conn, if_exists='replace', index=False)

# # Close the connection
# conn.close()